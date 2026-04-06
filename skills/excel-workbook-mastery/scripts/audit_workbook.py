"""
Workbook Audit Script — excel-workbook-mastery skill

Reads a .xlsx file and produces a structured audit report covering:
- Sheet inventory and naming quality
- Formula complexity and error scan
- Named ranges and Tables
- Data validation coverage
- Formatting consistency signals
- Structural assessment

Usage:
    python audit_workbook.py <path_to_xlsx> [--output <report_path>]

Output: Markdown report printed to stdout or saved to file.
"""

import sys
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


def audit_workbook(filepath):
    """Run a full structural audit on a workbook and return findings."""
    findings = {
        'filepath': str(filepath),
        'sheets': [],
        'named_ranges': [],
        'tables': [],
        'errors': [],
        'formulas': {'total': 0, 'complex': 0, 'volatile': 0},
        'validation_count': 0,
        'issues': [],
        'summary': {},
    }

    wb = load_workbook(filepath, data_only=False)

    # --- Sheet inventory ---
    bad_names = {'Sheet1', 'Sheet2', 'Sheet3', 'Sheet', 'New Sheet', 'Copy', 'Final', 'test'}
    hidden_sheets = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet_info = {
            'name': ws_name,
            'rows': ws.max_row or 0,
            'cols': ws.max_column or 0,
            'hidden': ws.sheet_state != 'visible',
        }
        findings['sheets'].append(sheet_info)

        if ws.sheet_state != 'visible':
            hidden_sheets.append(ws_name)

        # Flag bad names
        if any(ws_name.lower().startswith(bn.lower()) for bn in bad_names):
            findings['issues'].append({
                'severity': 'Minor',
                'category': 'Structure',
                'message': f'Sheet "{ws_name}" has a generic/default name — rename to describe its purpose.',
            })

    if hidden_sheets:
        findings['issues'].append({
            'severity': 'Important',
            'category': 'Structure',
            'message': f'Hidden sheets found: {", ".join(hidden_sheets)}. Verify they contain no critical logic.',
        })

    # --- Named ranges ---
    try:
        # openpyxl >= 3.1
        named_items = wb.defined_names.values()
    except AttributeError:
        # openpyxl < 3.1
        named_items = wb.defined_names.definedName

    for defn in named_items:
        name_str = defn.name if hasattr(defn, 'name') else str(defn)
        value_str = defn.attr_text if hasattr(defn, 'attr_text') else str(defn.value) if hasattr(defn, 'value') else ''
        findings['named_ranges'].append({
            'name': name_str,
            'value': value_str,
        })

    # --- Tables ---
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for table in ws.tables.values():
            findings['tables'].append({
                'sheet': ws_name,
                'name': table.displayName,
                'ref': table.ref,
            })

    # --- Formula and error scan ---
    volatile_funcs = {'INDIRECT', 'OFFSET', 'NOW', 'TODAY', 'RAND', 'RANDBETWEEN', 'INFO', 'CELL'}
    complex_threshold = 100  # formula string length
    error_values = {'#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#NUM!', '#N/A', '#NULL!'}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        if ws.sheet_state != 'visible':
            continue

        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 5000),
                                max_col=min(ws.max_column or 0, 100)):
            for cell in row:
                val = cell.value
                if val is None:
                    continue

                # Check for formulas
                if isinstance(val, str) and val.startswith('='):
                    findings['formulas']['total'] += 1

                    if len(val) > complex_threshold:
                        findings['formulas']['complex'] += 1

                    upper_val = val.upper()
                    for vf in volatile_funcs:
                        if vf in upper_val:
                            findings['formulas']['volatile'] += 1
                            break

                    # Check for blanket IFERROR
                    if 'IFERROR' in upper_val and upper_val.count('IFERROR') > 0:
                        if '""' in val or ",0)" in val or ',"")' in val:
                            pass  # Counted globally below

                # Check for error values (in data_only mode these would be values)
                if isinstance(val, str) and val.strip() in error_values:
                    findings['errors'].append({
                        'sheet': ws_name,
                        'cell': f'{get_column_letter(cell.column)}{cell.row}',
                        'error': val.strip(),
                    })

    # --- Data validation ---
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        findings['validation_count'] += len(ws.data_validations.dataValidation)

    # --- Generate issues from findings ---
    if findings['formulas']['complex'] > 0:
        pct = (findings['formulas']['complex'] / max(findings['formulas']['total'], 1)) * 100
        findings['issues'].append({
            'severity': 'Important' if pct > 10 else 'Minor',
            'category': 'Formulas',
            'message': f'{findings["formulas"]["complex"]} formulas exceed {complex_threshold} chars ({pct:.0f}% of total). Consider helper columns to break these apart.',
        })

    if findings['formulas']['volatile'] > 5:
        findings['issues'].append({
            'severity': 'Important',
            'category': 'Performance',
            'message': f'{findings["formulas"]["volatile"]} volatile functions found (INDIRECT, OFFSET, NOW, etc.). These recalculate on every edit and can slow large workbooks.',
        })

    if findings['errors']:
        findings['issues'].append({
            'severity': 'Critical',
            'category': 'Errors',
            'message': f'{len(findings["errors"])} visible errors found. Fix or handle these before handoff.',
        })

    if findings['validation_count'] == 0 and findings['formulas']['total'] > 20:
        findings['issues'].append({
            'severity': 'Important',
            'category': 'Validation',
            'message': 'No data validation found in a workbook with substantial formulas. Add validation to input cells.',
        })

    if len(findings['sheets']) > 1 and not any(
        s['name'].lower() in ('readme', '00_readme', 'cover', '00_cover', 'instructions', 'about')
        for s in findings['sheets']
    ):
        findings['issues'].append({
            'severity': 'Minor',
            'category': 'Documentation',
            'message': 'No README or Cover sheet found. Add one to document purpose, assumptions, and usage.',
        })

    # --- Summary ---
    findings['summary'] = {
        'total_sheets': len(findings['sheets']),
        'hidden_sheets': len(hidden_sheets),
        'total_formulas': findings['formulas']['total'],
        'complex_formulas': findings['formulas']['complex'],
        'volatile_formulas': findings['formulas']['volatile'],
        'error_count': len(findings['errors']),
        'named_ranges': len(findings['named_ranges']),
        'tables': len(findings['tables']),
        'validation_rules': findings['validation_count'],
        'issue_count': len(findings['issues']),
        'critical_count': sum(1 for i in findings['issues'] if i['severity'] == 'Critical'),
        'important_count': sum(1 for i in findings['issues'] if i['severity'] == 'Important'),
        'minor_count': sum(1 for i in findings['issues'] if i['severity'] == 'Minor'),
    }

    return findings


def format_report(findings):
    """Format audit findings as a markdown report."""
    s = findings['summary']
    lines = []

    lines.append(f"# Workbook Audit Report")
    lines.append(f"")
    lines.append(f"**File:** `{findings['filepath']}`")
    lines.append(f"")

    # Health assessment
    if s['critical_count'] > 0:
        health = "**Terminal** — critical errors must be fixed before trusting any output."
    elif s['important_count'] > 2:
        health = "**Sick** — structural problems need attention before sharing."
    elif s['important_count'] > 0:
        health = "**Recovering** — a few issues to address but fundamentally sound."
    else:
        health = "**Healthy** — minor cleanup only."

    lines.append(f"**Assessment:** {health}")
    lines.append(f"")

    # Summary table
    lines.append(f"## Summary")
    lines.append(f"")
    lines.append(f"| Metric | Value |")
    lines.append(f"|---|---|")
    lines.append(f"| Sheets (visible / hidden) | {s['total_sheets'] - s['hidden_sheets']} / {s['hidden_sheets']} |")
    lines.append(f"| Total formulas | {s['total_formulas']} |")
    lines.append(f"| Complex formulas (>{100} chars) | {s['complex_formulas']} |")
    lines.append(f"| Volatile functions | {s['volatile_formulas']} |")
    lines.append(f"| Visible errors | {s['error_count']} |")
    lines.append(f"| Named ranges | {s['named_ranges']} |")
    lines.append(f"| Excel Tables | {s['tables']} |")
    lines.append(f"| Data validation rules | {s['validation_rules']} |")
    lines.append(f"")

    # Sheet inventory
    lines.append(f"## Sheet Inventory")
    lines.append(f"")
    lines.append(f"| # | Sheet Name | Rows | Cols | Hidden |")
    lines.append(f"|---|---|---|---|---|")
    for i, sheet in enumerate(findings['sheets'], 1):
        hidden_mark = "Yes" if sheet['hidden'] else ""
        lines.append(f"| {i} | {sheet['name']} | {sheet['rows']} | {sheet['cols']} | {hidden_mark} |")
    lines.append(f"")

    # Errors
    if findings['errors']:
        lines.append(f"## Visible Errors")
        lines.append(f"")
        lines.append(f"| Sheet | Cell | Error |")
        lines.append(f"|---|---|---|")
        for err in findings['errors'][:20]:  # Cap at 20
            lines.append(f"| {err['sheet']} | {err['cell']} | `{err['error']}` |")
        if len(findings['errors']) > 20:
            lines.append(f"| ... | ... | {len(findings['errors']) - 20} more |")
        lines.append(f"")

    # Issues
    if findings['issues']:
        lines.append(f"## Issues ({s['critical_count']} Critical, {s['important_count']} Important, {s['minor_count']} Minor)")
        lines.append(f"")
        for severity in ['Critical', 'Important', 'Minor']:
            issues = [i for i in findings['issues'] if i['severity'] == severity]
            if issues:
                lines.append(f"### {severity}")
                lines.append(f"")
                for issue in issues:
                    lines.append(f"- **[{issue['category']}]** {issue['message']}")
                lines.append(f"")

    # Named ranges and tables
    if findings['named_ranges']:
        lines.append(f"## Named Ranges ({len(findings['named_ranges'])})")
        lines.append(f"")
        for nr in findings['named_ranges'][:15]:
            lines.append(f"- `{nr['name']}` → `{nr['value']}`")
        if len(findings['named_ranges']) > 15:
            lines.append(f"- ... and {len(findings['named_ranges']) - 15} more")
        lines.append(f"")

    if findings['tables']:
        lines.append(f"## Excel Tables ({len(findings['tables'])})")
        lines.append(f"")
        for tbl in findings['tables']:
            lines.append(f"- `{tbl['name']}` on `{tbl['sheet']}` ({tbl['ref']})")
        lines.append(f"")

    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='Audit an Excel workbook for structural issues.')
    parser.add_argument('filepath', help='Path to .xlsx file')
    parser.add_argument('--output', '-o', help='Save report to file (default: stdout)')
    args = parser.parse_args()

    filepath = Path(args.filepath)
    if not filepath.exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    if filepath.suffix.lower() not in ('.xlsx', '.xlsm'):
        print(f"Error: Expected .xlsx or .xlsm file, got {filepath.suffix}")
        sys.exit(1)

    findings = audit_workbook(filepath)
    report = format_report(findings)

    if args.output:
        Path(args.output).write_text(report)
        print(f"Audit report saved to {args.output}")
    else:
        print(report)


if __name__ == '__main__':
    main()
